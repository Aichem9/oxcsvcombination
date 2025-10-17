import io
import os
import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="엑셀 합치기 · Row7 시작", page_icon="🧩", layout="wide")

st.title("🧩 여러 엑셀 파일을 한 번에 이어붙이기")
\"\"\"
요구사항 반영:
1) **첫 번째 파일의 1~6행(양식)**을 그대로 유지해서 결과 엑셀의 맨 위에 배치합니다.
2) **첫 번째 파일에서 데이터가 끝난 다음 행(지표 설명)**을 병합 결과의 **맨 마지막 행**으로 보냅니다.

- 기본 가정: 헤더는 7행(사람 기준), 데이터는 8행부터 시작.
- 각 파일의 마지막 유효 행은 자동 감지.
- 완전히 빈 열/행 제거, 중복 열 이름 정규화.
\"\"\"

# --- 사이드바 옵션 ---
st.sidebar.header("옵션")
start_row_human = st.sidebar.number_input(
    "헤더(열 이름)가 위치한 행 번호", min_value=1, value=7, step=1,
    help="사람 기준 행 번호입니다. 기본값 7행."
)
add_filename_col = st.sidebar.checkbox("원본 파일명 열 추가", value=True)

st.sidebar.markdown("---")
st.sidebar.caption("파일은 세션 내 임시 처리되며 서버에 저장되지 않습니다.")

# --- 유틸 함수(표 병합용) ---
def read_and_trim_excel(file, start_row_header_idx: int) -> pd.DataFrame:
    df_all = pd.read_excel(file, header=None)
    if df_all.empty:
        return pd.DataFrame()

    header_series = df_all.iloc[start_row_header_idx].astype(object)
    header_series = header_series.where(~header_series.isna(), None).ffill()
    header = header_series.tolist()

    body = df_all.iloc[start_row_header_idx + 1 : ].copy()
    if body.dropna(how="all").empty:
        trimmed = pd.DataFrame(columns=header)
    else:
        last_idx = body.dropna(how="all").index[-1]
        trimmed = body.loc[:last_idx].copy()
        trimmed.columns = header
        trimmed = trimmed.dropna(how="all")
        trimmed = trimmed.dropna(axis=1, how="all")
        counts = {}
        new_cols = []
        for c in trimmed.columns:
            c = str(c)
            if c in counts:
                counts[c] += 1
                new_cols.append(f"{c}.{counts[c]}")
            else:
                counts[c] = 0
                new_cols.append(c)
        trimmed.columns = new_cols
    return trimmed, df_all

def detect_description_row_index(df_all: pd.DataFrame, start_row_header_idx: int):
    \"\"\"첫 파일에서만 사용: 데이터 마지막 다음 행(지표 설명) 인덱스(0-based)를 찾는다.\n
    - 데이터 본문은 start_row_header_idx+1 부터 시작\n
    - 본문에서 마지막 유효 행을 찾은 뒤 +1 한 행이 설명 행으로 간주\n
    - 해당 행이 파일 범위를 넘으면 None 반환\n    \"\"\"\n    body = df_all.iloc[start_row_header_idx + 1 : ]
    if body.dropna(how=\"all\").empty:\n        return None\n    last_idx = body.dropna(how=\"all\").index[-1]\n    desc_idx = last_idx + 1\n    if desc_idx < len(df_all):\n        return desc_idx\n    return None\n
def copy_first_six_rows(src_wb, dst_wb):\n    \"\"\"openpyxl 워크북 간 1~6행 셀 값과 병합 범위를 복제(서식은 최소한).\"\"\"\n    src_ws = src_wb.active\n    dst_ws = dst_wb.active\n\n    max_col = src_ws.max_column\n    # 값 복사 (1~6행)\n    for r in range(1, 7):\n        for c in range(1, max_col + 1):\n            dst_ws.cell(row=r, column=c, value=src_ws.cell(row=r, column=c).value)\n\n    # 병합 복사 (1~6행에 걸친 범위만)\n    for mr in src_ws.merged_cells.ranges:\n        min_row, min_col, max_row, max_col = mr.min_row, mr.min_col, mr.max_row, mr.max_col\n        if max_row <= 6:  # 1~6행 범위 내 병합만 반영\n            dst_ws.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)\n\n    # 열 너비 복사(가독성용)\n    for col_dim in src_ws.column_dimensions:\n        width = src_ws.column_dimensions[col_dim].width\n        if width:\n            dst_ws.column_dimensions[col_dim].width = width\n\nuploaded_files = st.file_uploader(\n    \"같은 양식의 엑셀 파일을 선택하세요 (여러 개 가능)\",\n    type=[\"xlsx\", \"xls\"], accept_multiple_files=True\n)\n\nif uploaded_files:\n    st.write(f\"선택한 파일 수: **{len(uploaded_files)}**\")\n\n    if st.button(\"파일 합치기 실행\", type=\"primary\"):\n        dataframes = []\n        errors = []\n        first_file_raw_df = None\n        first_file_stream_copy = None  # openpyxl용 원본 보관\n        desc_row_values = None\n\n        for idx, f in enumerate(uploaded_files):\n            try:\n                # pandas가 읽으면 스트림 포인터가 이동할 수 있어 복사용 버퍼를 따로 둠\n                raw_bytes = f.read()\n                f.seek(0)\n                df, df_all = read_and_trim_excel(f, start_row_header_idx=int(start_row_human - 1))\n\n                if add_filename_col:\n                    fname = getattr(f, \"name\", \"uploaded.xlsx\")\n                    df.insert(0, \"source_file\", os.path.basename(fname))\n                dataframes.append(df)\n\n                if idx == 0:\n                    first_file_raw_df = df_all  # 첫 파일 전체 원시 DF\n                    first_file_stream_copy = io.BytesIO(raw_bytes)  # 원본 바이너리(copy)\n                    # 지표 설명 행 탐지(0-based index)\n                    desc_idx = detect_description_row_index(df_all, int(start_row_human - 1))\n                    if desc_idx is not None:\n                        # 해당 행의 모든 값을 리스트로 저장\n                        desc_row_values = df_all.iloc[desc_idx].tolist()\n            except Exception as e:\n                errors.append((getattr(f, \"name\", \"(이름 없음)\"), str(e)))\n\n        if errors:\n            with st.expander(\"읽기 오류가 있었습니다 (펼치기)\"):\n                for name, msg in errors:\n                    st.error(f\"{name}: {msg}\")\n\n        if dataframes:\n            merged = pd.concat(dataframes, ignore_index=True)\n            st.success(f\"병합 완료! 행 {merged.shape[0]} · 열 {merged.shape[1]}\")\n            st.dataframe(merged.head(200), use_container_width=True)\n\n            # === 엑셀 내보내기(첫 파일 1~6행 보존 + 헤더 7행 + 데이터 + 마지막에 설명 행) ===\n            out_buf = io.BytesIO()\n\n            # 1) 새 워크북 생성 및 첫 6행 템플릿 복사\n            wb_out = Workbook()\n            ws_out = wb_out.active\n\n            if first_file_stream_copy is not None:\n                try:\n                    wb_src = load_workbook(first_file_stream_copy, data_only=True)\n                    copy_first_six_rows(wb_src, wb_out)\n                except Exception:\n                    pass  # 템플릿 복사 실패해도 계속 진행\n\n            # 2) pandas로 병합 표를 7행부터 쓰기(startrow=6)\n            from openpyxl.utils.dataframe import dataframe_to_rows\n            # 헤더와 데이터 모두 작성하도록 직접 rows를 사용\n            start_row = 7  # 사람 기준 7행\n            # 헤더\n            for j, col_name in enumerate(merged.columns, start=1):\n                ws_out.cell(row=start_row, column=j, value=str(col_name))\n            # 데이터\n            for i in range(len(merged)):\n                for j, col_name in enumerate(merged.columns, start=1):\n                    ws_out.cell(row=start_row + i + 1, column=j, value=merged.iat[i, j-1])\n\n            # 3) 맨 아래에 '지표 설명' 행 추가(첫 파일에서 추출)\n            if desc_row_values is not None:\n                append_row = start_row + 1 + len(merged)\n                for j, val in enumerate(desc_row_values, start=1):\n                    ws_out.cell(row=append_row, column=j, value=val)\n\n            wb_out.save(out_buf)\n            out_buf.seek(0)\n\n            # CSV도 함께 제공(템플릿/설명 행 특성상 CSV에는 템플릿 보존이 어렵지만 데이터는 동일)\n            csv_buf = io.BytesIO()\n            csv_buf.write(merged.to_csv(index=False, encoding=\"utf-8-sig\").encode(\"utf-8-sig\"))\n            csv_buf.seek(0)\n\n            col1, col2 = st.columns(2)\n            with col1:\n                st.download_button(\n                    \"엑셀 파일로 다운로드 (.xlsx) — 템플릿 상단 유지 & 설명 행 하단 이동\",\n                    data=out_buf,\n                    file_name=\"merged_with_template_and_desc.xlsx\",\n                    mime=\"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\",\n                    use_container_width=True,\n                )\n            with col2:\n                st.download_button(\n                    \"CSV로 다운로드 (.csv)\",\n                    data=csv_buf,\n                    file_name=\"merged_output.csv\",\n                    mime=\"text/csv\",\n                    use_container_width=True,\n                )\n        else:\n            st.warning(\"유효한 데이터가 없어요. 헤더 행 번호가 맞는지 확인해주세요.\")\nelse:\n    st.info(\"상단에서 엑셀 파일을 올려주세요.\")\n